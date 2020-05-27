#!/usr/bin/python
# -*- coding: utf-8 -*-
#File Name: excel_arxml_file_creation.py

#**************************************************************************************************************
#*  AUTHOR IDENTITY:
#*-------------------------------------------------------------------------------------------------------------
#*  Full Name:                        Last Name                            Company
#*-------------------------------------------------------------------------------------------------------------
#*  Mac Tien Truong                     Truong                             Hitachi
#
#*  REVISION HISTORY
#*-------------------------------------------------------------------------------------------------------------
#*  Verion                            Date            Auhor               Description
#*-------------------------------------------------------------------------------------------------------------
#*  1.0.                             2019-10-08      Truong             Create arxml file.
#*
#**************************************************************************************************************/

import os
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, GradientFill
from openpyxl.styles import colors
from openpyxl.cell import Cell
from lib import*


#============================================== Create an new excel file ======================================
#==============================================================================================================
def createNewExcelFile (excel_file_name):
    """This function is used to create an excel file with its worlksheet"""

    #Define the local variable(s) here
    l_excel_file_name = excel_file_name

    l_workbook = openpyxl.Workbook()
    l_workbook.save(l_excel_file_name)

    return()

#======================================== Create a worksheet in selected file =================================
#==============================================================================================================
def addNewWorkSheet (excel_file_name, worksheet):
    """This function is used to create new worksheet in provided file if it is not existing"""

    #Define the local variable(s) here
    l_excel_file_name = excel_file_name
    l_worksheet = worksheet

    #Load the provided file
    l_workbook = openpyxl.load_workbook(l_excel_file_name)

    #Get all the worksheets contained in this file
    l_current_worksheet = l_workbook.get_sheet_names()

    #Check if worksheet is not existing then create it else do nothing
    l_found = "No"
    for l_each_current_worksheet in l_current_worksheet:
        if(l_each_current_worksheet == l_worksheet):
            l_found = "Yes"
            break

    #If this worksheet is not existing then create it
    if(l_found == "No"):
        l_workbook.create_sheet(l_worksheet)

    #Save file
    l_workbook.save(l_excel_file_name)

    return()

#======================================== Remove a worksheet in selected file =================================
#==============================================================================================================
def removeWorkSheet (excel_file_name, worksheet):
    """This function is used to remove provided worksheet in selected file if it is existing"""

    #Define the local variable(s) here
    l_excel_file_name = excel_file_name
    l_worksheet = worksheet

    #Load the provided file
    l_workbook = openpyxl.load_workbook(l_excel_file_name)

    #Get all the worksheets contained in this file
    l_current_worksheet = l_workbook.get_sheet_names()

    #Check if worksheet is existing then remove it else do nothing
    l_found = "No"
    for l_each_current_worksheet in l_current_worksheet:
        if(l_each_current_worksheet == l_worksheet):
            l_found = "Yes"
            break

    #If this worksheet is existing then remove it
    if(l_found == "Yes"):
        l_std = l_workbook.get_sheet_by_name(l_worksheet)
        l_workbook.remove_sheet(l_std)

    #Save file
    l_workbook.save(l_excel_file_name)

    return()

#========================================== Write data in a selected cell =====================================
#==============================================================================================================
def writeDataBycell(excel_file_name, worksheet, row, column, data):
    """This function is used to write data in a selected cell. If file or worksheet is 
        not existing then create new"""

    #Define the local variable(s) here
    l_excel_file_name = excel_file_name
    l_worksheet = worksheet
    l_start_row = row
    l_start_column = column
    l_data = data

    #If file is not existing then create new else open it
    if(not os.path.exists(l_excel_file_name)):
        l_workbook = openpyxl.Workbook()
        l_workbook.save(l_excel_file_name)

    #Create worksheet if it not existing
    addNewWorkSheet(l_excel_file_name, l_worksheet)
    #Remove worksheet "Sheet" if it is existing
    removeWorkSheet(l_excel_file_name, "Sheet")

    l_workbook = openpyxl.load_workbook(l_excel_file_name)
    #Get worksheet that will be written dara to
    l_ws = l_workbook.get_sheet_by_name(l_worksheet)

    #Write data into excel file
    l_cell = l_ws.cell(row = l_start_row, column = l_start_column)
    l_cell.value = l_data

    #Save file
    l_workbook.save(l_excel_file_name)

    return()

#=============================================== Write data by a data table ===================================
#==============================================================================================================
def writeDataByDataTable(excel_file_name, worksheet, start_row, start_column, data_table):
    """This function is used to write data by a data table into a worksheet of an excel file"""

    #Define the local variable(s) here
    l_excel_file_name = excel_file_name
    l_worksheet = worksheet
    l_start_row = start_row
    l_start_column = start_column
    l_data_table = data_table

    #If file is not existing then create new else open it
    if(not os.path.exists(l_excel_file_name)):
        l_workbook = openpyxl.Workbook()
        l_workbook.save(l_excel_file_name)

    #Create worksheet if it not existing
    addNewWorkSheet(l_excel_file_name, l_worksheet)
    #Remove worksheet "Sheet" if it is existing
    removeWorkSheet(l_excel_file_name, "Sheet")

    l_workbook = openpyxl.load_workbook(l_excel_file_name)
    #Get worksheet that will be written dara to
    l_ws = l_workbook.get_sheet_by_name(l_worksheet)
    #l_ws = l_workbook.active

    #Write data into excel file
    for l_each_column in range(len(l_data_table)):
        for l_each_row in range(len(l_data_table[l_each_column])):
            l_cell = l_ws.cell(row = l_each_row + l_start_row, \
                     column = l_each_column + l_start_column)
            l_cell.value = l_data_table[l_each_column][l_each_row]

    #Save file
    l_workbook.save(l_excel_file_name)

    return()


#=============================================== Write data by a data table ===================================
#==============================================================================================================
def writeDataByDataTableByAddingFormat(excel_file_name, worksheet, start_row, start_column, data_table, ptoject_name, Sip_name, information):
    """This function is used to write data by a data table into a worksheet of an excel file"""

    #Define the local variable(s) here
    l_excel_file_name = excel_file_name
    l_worksheet = worksheet
    l_start_row = start_row
    l_start_column = start_column
    l_data_table = data_table
    l_ptoject_name = ptoject_name
    l_Sip_name = Sip_name
    l_infor = information

    l_latest_col_index = len(l_data_table) + l_start_column - 1
    l_latest_row_index = len(l_data_table[0]) + l_start_row - 1


    #Border define
    l_medium = Side(border_style="medium", color="000000")
    l_thin = Side(border_style="thin", color="000000")

    #Color define
    l_color_index = [50, 5, 4, 6, 7, 3, 29, 40, 53, 52, 47, 45, 43, 27, 45]
    colorfill = []
    for each_color_index in l_color_index:
        colorfill.append( PatternFill(start_color=colors.COLOR_INDEX[each_color_index],
                                      end_color=colors.COLOR_INDEX[each_color_index],
                                      fill_type='solid'))

    #If file is not existing then create new else open it
    if(not os.path.exists(l_excel_file_name)):
        l_workbook = openpyxl.Workbook()
        l_workbook.save(l_excel_file_name)

    print("Start creating the file: ", ptoject_name + "/"+ Sip_name + "/" + excel_file_name)
    #Create worksheet if it not existing and remove the sheet name "Sheet" if it is existing
    addNewWorkSheet(l_excel_file_name, l_worksheet)
    removeWorkSheet(l_excel_file_name, "Sheet")

    #Load workbook and get worksheet that will be written dara to
    l_workbook = openpyxl.load_workbook(l_excel_file_name)
    l_ws = l_workbook.get_sheet_by_name(l_worksheet)

    #Write data into excel file
    for l_each_column in range(len(l_data_table)):
        for l_each_row in range(len(l_data_table[l_each_column])):
            l_cell = l_ws.cell(row = l_each_row + l_start_row, \
                               column = l_each_column + l_start_column)
            l_cell.value = l_data_table[l_each_column][l_each_row]
            l_cell.alignment = Alignment(horizontal="left", vertical="top")
            l_row_count = l_each_row + l_start_row
            #Process for columns in the Main and sub
            if((l_row_count != l_start_row) and (l_each_column  + l_start_column < l_latest_col_index - 4)):
                #For the cell contains data
                if(len(l_data_table[l_each_column][l_each_row]) != 0):
                    l_col_count = l_each_column + l_start_column
                    #For the row that is not empty. border and set corlor
                    while (l_col_count < l_latest_col_index):
                        l_ws.cell(row = l_row_count, column = l_col_count).fill = colorfill[l_each_column + 1]
                        l_ws.cell(row = l_row_count, column = l_col_count).border = Border(top=l_thin)
                        if (l_row_count == l_latest_row_index):
                            l_ws.cell(row = l_latest_row_index, column = l_col_count).border = Border(top=l_thin, bottom=l_medium)
                        l_col_count = l_col_count + 1

                    #Border for this cell
                    l_col_count = l_each_column + l_start_column
                    #For last right cell
                    if(l_col_count == l_start_column):
                        l_ws.cell(row = l_row_count, column = l_col_count).border = Border(top=l_thin, left=l_medium)
                        if(l_row_count == l_latest_row_index):
                            l_ws.cell(row = l_latest_row_index, column = l_col_count).border = Border(top=l_thin, left=l_medium, bottom=l_medium)
                    #For between cells
                    else:
                        l_ws.cell(row = l_row_count, column = l_col_count).border = Border(top=l_thin, left=l_thin)
                        if(l_row_count == l_latest_row_index):
                            l_ws.cell(row = l_latest_row_index, column = l_col_count).border = Border(top=l_thin, left=l_thin, bottom=l_medium)
                    #For last left cell
                    if(l_row_count == l_latest_row_index):
                        l_ws.cell(row = l_row_count, column = l_latest_col_index - 1).border = Border(top=l_thin, right=l_thin, bottom=l_medium)
                    else:
                        l_ws.cell(row = l_row_count, column = l_latest_col_index - 1).border = Border(top=l_thin, right=l_thin)

                    #For next empty row in first column
                    if (l_start_column == l_col_count):
                        if (l_each_row < len(l_data_table[l_each_column]) - 1):
                            l_row_next_index = 1
                            l_found = False
                            while((False == l_found)and ((l_each_row + l_row_next_index) < len(l_data_table[l_each_column]))):
                                if(l_data_table[l_each_column][l_each_row + l_row_next_index] != ""):
                                    l_found = True
                                if (False == l_found):
                                    l_ws.cell(row = l_row_count + l_row_next_index, column = l_col_count).border = Border(left=l_medium)
                                    l_ws.cell(row = l_row_count + l_row_next_index, column = l_col_count).fill = colorfill[l_each_column + 1]
                                    if (l_row_count + l_row_next_index == l_latest_row_index):
                                        l_ws.cell(row = l_latest_row_index, column = l_col_count).border = Border(left=l_medium, bottom=l_medium)
                                        l_ws.cell(row = l_latest_row_index, column = l_col_count).fill = colorfill[l_each_column + 1]
                                l_row_next_index = l_row_next_index + 1

                    #For next empty row in next columns
                    else:
                        if(l_each_row < len(l_data_table[l_each_column]) - 1):
                            l_col_count = l_each_column + l_start_column
                            l_row_next_index = 1
                            f_found = False
                            while(((l_each_row + l_row_next_index) < len(l_data_table[l_each_column])) and (f_found == False)):
                                l_col_index = l_col_count
                                while (l_col_index >=  l_start_column):
                                    if(l_data_table[l_col_index - l_start_column][l_each_row + l_row_next_index] != ""):
                                        f_found = True
                                        break
                                    l_col_index = l_col_index - 1
                                if(f_found == False):
                                    l_row_next_index = l_row_next_index + 1
                            l_row_next_index = l_row_next_index - 1
                            if (l_row_next_index > 0):
                                #Fill color and border for next empty row
                                for l_each in range(l_row_next_index):
                                    l_ws.cell(row = l_row_count + l_each + 1, column = l_col_count).border = Border(left=l_thin)
                                    l_ws.cell(row = l_row_count + l_each + 1, column = l_col_count).fill = colorfill[l_each_column + 1]
                                    if (l_row_count + l_each + 1 == l_latest_row_index):
                                        l_ws.cell(row = l_latest_row_index, column = l_col_count).fill = colorfill[l_each_column + 1]
                                        l_ws.cell(row = l_latest_row_index, column = l_col_count).border = Border(left=l_thin, bottom=l_medium)
                                l_ws.cell(row = l_row_count, column = l_col_count).border = Border(left=l_thin, top=l_thin)

                    #For its paremeter rows
                    l_col_count = l_each_column + l_start_column
                    l_row_next_index = 1
                    f_found = False
                    while(((l_each_row + l_row_next_index) < len(l_data_table[l_each_column])) and (f_found == False)):
                        l_col_index = l_col_count + 1
                        while (l_col_index >=  l_start_column):
                            if(l_data_table[l_col_index - l_start_column][l_each_row + l_row_next_index] != ""):
                                f_found = True
                                break
                            l_col_index = l_col_index - 1
                        if(f_found == False):
                            l_row_next_index = l_row_next_index + 1

                    l_row_next_index = l_row_next_index - 1
                    if (l_row_next_index > 0):
                        #Fill color and border for next empty row
                        for l_each in range(l_row_next_index):
                            l_col_next = l_col_count + 1
                            while (l_col_next < l_latest_col_index - 4):
                                l_ws.cell(row = l_row_count + l_each + 1, column = l_col_next).fill = colorfill[l_each_column + 1]
                                if (l_row_count + l_each + 1 == l_latest_row_index):
                                    l_ws.cell(row = l_latest_row_index, column = l_col_next).fill = colorfill[l_each_column + 1]
                                    l_ws.cell(row = l_latest_row_index, column = l_col_next).border = Border(bottom=l_medium)
                                l_col_next = l_col_next + 1

            #For columns of parameters
            elif(l_each_column  + l_start_column < l_latest_col_index):
                if (l_each_row < len(l_data_table[l_each_column]) - 1):
                    if(l_data_table[l_latest_col_index - 5][l_each_row] != ""):
                        l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_thin)
                elif(l_each_row == len(l_data_table[l_each_column]) - 1):
                    if(l_data_table[l_latest_col_index - 5][l_each_row] != ""):
                        l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_medium)
                if (l_each_column + l_start_column == l_latest_col_index - 2):
                    l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif(l_each_column  + l_start_column == l_latest_col_index):
                if (l_each_row < len(l_data_table[l_each_column]) - 1):
                    if(l_data_table[l_each_column][l_each_row] != ""):
                        l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).border = Border(top=l_thin, left=l_thin, right=l_medium, bottom=l_thin)
                        l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                elif(l_each_row == len(l_data_table[l_each_column]) - 1):
                    if(l_data_table[l_each_column][l_each_row] != ""):
                        l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).border = Border(top=l_thin, left=l_thin, right=l_medium, bottom=l_medium)
                        l_ws.cell(row = l_each_row + l_start_row, column = l_each_column + l_start_column).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            l_ws.cell(row = l_latest_row_index, column = l_latest_col_index).border = Border(top=l_thin, left=l_thin, right=l_medium, bottom=l_medium)

    #Adjust some columns width
    l_ws.column_dimensions[str(chr(64 + 1))].width = 3*1.2
    l_ws.column_dimensions[str(chr(64 + l_latest_col_index))].width = 50*1.2
    l_ws.column_dimensions[str(chr(64 + l_latest_col_index - 1))].width = 15*1.2
    l_count = l_latest_col_index - 2
    while (l_count > l_latest_col_index - 5):
        l_ws.column_dimensions[str(chr(64 + l_count))].width = 27*1.2
        l_count = l_count - 1

    #Container Name cell
    l_ws.cell(row = l_start_row - 1, column = start_column).value = "Container Name"
    l_ws.merge_cells(start_row = l_start_row -1 , start_column = l_start_column, end_row = l_start_row - 1, end_column = l_latest_col_index - 5)
    l_ws.cell(row = l_start_row - 1, column = start_column).alignment = Alignment(horizontal="center", vertical="center")
    l_ws.cell(row = l_start_row - 1, column = start_column).border = Border(top=l_medium, left=l_medium, right=l_medium, bottom=l_medium)
    l_ws.cell(row = l_start_row - 1, column = start_column).font = Font(bold = True)
    l_ws.cell(row = l_start_row - 1, column = start_column).fill = colorfill[0]

    #Ptoject tile
    l_ws.cell(row = 2, column = l_latest_col_index - 4).value = l_infor
    l_ws.cell(row = 2, column = l_latest_col_index - 4).alignment = Alignment(horizontal="left", vertical="top")
    l_ws.cell(row = 2, column = l_latest_col_index - 4).font = Font(bold = True)
    l_ws.cell(row = 2, column = l_latest_col_index - 4).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_thin)
    l_ws.cell(row = 2, column = l_latest_col_index - 4).fill = colorfill[0]
    
    l_ws.cell(row = 3, column = l_latest_col_index - 4).value = "Project Name:"
    l_ws.cell(row = 3, column = l_latest_col_index - 4).alignment = Alignment(horizontal="right", vertical="top")
    l_ws.cell(row = 3, column = l_latest_col_index - 4).font = Font(bold = True)
    l_ws.cell(row = 3, column = l_latest_col_index - 4).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_thin)
    l_ws.cell(row = 3, column = l_latest_col_index - 3).value = l_ptoject_name
    l_ws.cell(row = 3, column = l_latest_col_index - 3).alignment = Alignment(horizontal="left", vertical="top")
    l_ws.cell(row = 3, column = l_latest_col_index - 3).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_thin)

    l_ws.cell(row = 4, column = l_latest_col_index - 4).value = "SIP Name:"
    l_ws.cell(row = 4, column = l_latest_col_index - 4).alignment = Alignment(horizontal="right", vertical="top")
    l_ws.cell(row = 4, column = l_latest_col_index - 4).font = Font(bold = True)
    l_ws.cell(row = 4, column = l_latest_col_index - 4).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_thin)
    l_ws.cell(row = 4, column = l_latest_col_index - 3).value = l_Sip_name
    l_ws.cell(row = 4, column = l_latest_col_index - 3).alignment = Alignment(horizontal="left", vertical="top")
    l_ws.cell(row = 4, column = l_latest_col_index - 3).border = Border(top=l_thin, left=l_thin, right=l_thin, bottom=l_thin)

    #Merge the sub cell
    l_ws.merge_cells(start_row = l_start_row, start_column = l_start_column + 1, end_row = l_start_row, end_column = l_latest_col_index - 5)
    #Alignment and border the column title
    l_count = l_start_column
    while (l_count <= l_latest_col_index):
        l_ws.cell(row = l_start_row, column = l_count).alignment = Alignment(horizontal="center", vertical="center")
        if(l_count == l_start_column):
            l_ws.cell(row = l_start_row, column = l_count).border = Border(top=l_medium, left=l_medium, right=l_thin, bottom=l_thin)
        elif (l_count == l_latest_col_index):
            l_ws.cell(row = l_start_row, column = l_count).border = Border(top=l_medium, left=l_thin, right=l_medium, bottom=l_thin)
        else:
            l_ws.cell(row = l_start_row, column = l_count).border = Border(top=l_medium, left=l_thin, right=l_thin, bottom=l_thin)
        if (l_count <= l_latest_col_index - 5):
            l_ws.cell(row = l_start_row - 1, column = l_count).border = Border(top=l_medium, left=l_medium, right=l_medium, bottom=l_medium)
        l_ws.cell(row = l_start_row, column = l_count).font = Font(bold = True)
        l_ws.cell(row = l_start_row, column = l_count).fill = colorfill[0]
        l_count = l_count + 1

    
    #Hide the gridlines in excel
    l_ws.sheet_view.showGridLines = False
    #Save file
    l_workbook.save(l_excel_file_name)
    print("Finish creating the file: ", ptoject_name + "/" + Sip_name + "/" + excel_file_name)
    return()

#============================================== Find maximum of deep nodes  ===================================
#==============================================================================================================
def findMaxDeepNumber(container_object, maxdeep):
    """This function is used to find out the maxixmum munber of node deep """

    #Define local variables here
    l_max_deep = maxdeep
    l_container_object = container_object

    for l_each_ecu_obj in l_container_object.ecuc_container_list:
        if(l_max_deep[0] < l_each_ecu_obj.ecu_node_deep):
            l_max_deep[0] = l_each_ecu_obj.ecu_node_deep
        for l_each_parameter in l_each_ecu_obj.parameters_list:
            if (l_max_deep[0] < l_each_parameter.parameter_deep):
                l_max_deep[0] = l_each_parameter.parameter_deep

        for l_each_sub_container in l_each_ecu_obj.sub_container_list:
            findMaxDeepNumber(l_each_sub_container, l_max_deep)

    return()

#=============================================== Fill data in an array  =======================================
#==============================================================================================================
def fillDataTOArray(container_object, data_array, array_lengh):
    """This function is used to fill data from Container Objest to an 2-dimensional  arrays"""

    #Define local variables here
    l_container_object = container_object
    l_data_array = data_array
    l_array_lengh = array_lengh

    for l_each_ecu_obj in l_container_object.ecuc_container_list:
        if(l_each_ecu_obj != "NULL"):
            l_ecu_node_deep = int(l_each_ecu_obj.ecu_node_deep)
            for l_each in range(l_array_lengh - 1):
                if(int(l_each) == l_ecu_node_deep):
                    l_data_array[l_each].append(str(l_each_ecu_obj.ecu_container_name))
                else:
                    l_data_array[l_each].append("")
            l_data_array[len(l_data_array) - 1].append(str(l_each_ecu_obj.node_Description))
            for l_each_parameter in l_each_ecu_obj.parameters_list:
                for l_each in range(l_array_lengh):
                    if(l_each < l_array_lengh - 5):
                        l_data_array[l_each].append("")
                    elif(l_each == l_array_lengh - 5):
                        l_data_array[l_each].append(str(l_each_parameter.parameter_name))
                    elif(l_each == l_array_lengh - 4):
                        l_data_array[l_each].append(str(l_each_parameter.parameter_type))
                    elif(l_each == l_array_lengh - 3):
                        if(l_each_parameter.parameter_max != "" ) or (l_each_parameter.parameter_min != "" ):
                            l_range = str(l_each_parameter.parameter_min) + ".." + str(l_each_parameter.parameter_max)
                        elif (l_each_parameter.parameter_type == "EcucBooleanParamDef"):
                            l_range = "TRUE/FALSE"
                        else:
                            l_range = ""
                            for l_each_value in l_each_parameter.parameter_list_value:
                                l_range = l_range + str(l_each_value) + " \n"
                        l_data_array[l_each].append(str(l_range))
                    elif(l_each == l_array_lengh - 2):
                        l_data_array[l_each].append(str(l_each_parameter.parameter_default))
                    elif(l_each == l_array_lengh - 1):
                        l_data_array[l_each].append(str(l_each_parameter.parameter_description))

            for l_each_sub_container in l_each_ecu_obj.sub_container_list:
                if(l_each_sub_container != "NULL"):
                    fillDataTOArray(l_each_sub_container, data_array, array_lengh)

    return(True)

#======================================== Write data to output file information ===============================
#==============================================================================================================
def fillDataToOutputExcelFile(container_object, output_file, sheet_name, project_name, sip_name, information):
    """This function is used to write rawdata into output files"""

    #Define local variables here
    l_container_object = container_object
    l_output_file = output_file
    l_start_row = 8
    l_start_col = 2
    l_sheet_name = sheet_name
    l_project_name = project_name
    l_sip_name = sip_name
    l_infor = information

    #Find the node that is deepest in the arxml file
    l_max_deep = []
    l_max_deep.append(l_container_object.container_deep)
    findMaxDeepNumber(l_container_object, l_max_deep)
    l_length_array = l_max_deep[0] + 11

    #Using this 2-dimensional arrays to store all contrainer and its value
    l_data_array = [["Main"]]
    #Mark sub-container level
    for l_each in range(l_length_array):
        if(l_each != 0):
            if (l_each == 1):
                l_sub_container = "Sub"
            elif (l_each < l_length_array - 5):
                l_sub_container = ""
            elif (l_each == l_length_array - 5):
                l_sub_container = "Parametter/Reference Name"
            elif (l_each == l_length_array - 4):
                l_sub_container = "Type"
            elif (l_each == l_length_array - 3):
                l_sub_container = "Range/Reference to"
            elif (l_each == l_length_array - 2):
                l_sub_container = "Default value"
            elif (l_each == l_length_array - 1):
                l_sub_container = "Description"
            l_data_array.append([l_sub_container])


    fillDataTOArray(l_container_object, l_data_array, l_length_array)
    #Fill data into excel file
    writeDataByDataTableByAddingFormat(l_output_file, \
                                       l_sheet_name, \
                                       l_start_row, \
                                       l_start_col, \
                                       l_data_array, \
                                       l_project_name, \
                                       l_sip_name, \
                                       l_infor
                                      )

    return()